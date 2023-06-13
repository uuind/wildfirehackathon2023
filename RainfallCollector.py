from openpyxl.reader.excel import load_workbook
import xml.etree.ElementTree as ET

file = "24HrRainfallData.xlsx"

wb = load_workbook(filename = file, data_only=True)
ws = wb.active

landslides = tuple(ws.rows)

for row_index in range(1, len(landslides)):
    entry = landslides[row_index]
    if str(entry[13].value) == "yes":
        event_id = str(entry[7].value)
        rainfile = "Rainfall_Data/" + event_id + ".xml"
        tree = ET.parse(rainfile)
        root = tree.getroot()
        if len(root[0][0]) != 0:
            if isinstance(root[0][0][0][0].text, type(None)):
                day0 = 0
            else:
                day0 = float(root[0][0][0][0].text)

            if isinstance(root[0][0][1][0].text, type(None)):
                day1 = 0
            else:
                day1 = float(root[0][0][1][0].text)
        else:
            day0 = 0
            day1 = 0
        ws['O' + str(row_index + 1)] = day0
        ws['P' + str(row_index + 1)] = day1

wb.save("modified_rainfall.xlsx")