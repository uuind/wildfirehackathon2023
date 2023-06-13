from openpyxl.reader.excel import load_workbook

file = "fullynegativerainfall.xlsx"

wb = load_workbook(filename = file, data_only=True)
ws = wb.active
ws2 = wb.create_sheet("split")

splitting_factor = input("Every nth number will be selected for the dataset")
copy_index = 1
data = tuple(ws.rows)
for row_index in range(0, len(data), int(splitting_factor)):
    entry = data[row_index]
    ws2['A' + str(copy_index)] = entry[0].value
    ws2['B' + str(copy_index)] = entry[1].value
    ws2['C' + str(copy_index)] = entry[2].value

    copy_index += 1

wb.save("modified2" + file)