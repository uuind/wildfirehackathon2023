from openpyxl.reader.excel import load_workbook

file = "modifiedrainfall_negative.xlsx.xlsx"

wb = load_workbook(filename = file, data_only=True)
ws = wb.active
ws2 = wb.create_sheet("split")

splitting_factor = input("Every nth number will be selected for the dataset")
copy_index = 1
data = tuple(ws.rows)
for row_index in range(1, len(data), int(splitting_factor)):
    entry = data[row_index]
    ws2['A' + str(copy_index)] = entry[10].value
    ws2['B' + str(copy_index)] = entry[11].value
    ws2['C' + str(copy_index)] = entry[12].value
    ws2['D' + str(copy_index)] = entry[13].value
    ws2['E' + str(copy_index)] = entry[19].value
    ws2['F' + str(copy_index)] = entry[20].value
    ws2['G' + str(copy_index)] = entry[21].value
    ws2['H' + str(copy_index)] = entry[22].value

    copy_index += 1

wb.save("modified" + file)