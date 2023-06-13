from openpyxl.reader.excel import load_workbook
import xml.etree.ElementTree as ET

file = "rainfall_predictive.xlsx"

wb = load_workbook(filename = file, data_only=True)
ws = wb.active

landslides = tuple(ws.rows)

for row_index in range(1, len(landslides)):
    entry = landslides[row_index]
    station_name = str(entry[5].value)
    day = entry[4].value
    rainfile = "Rainfall_Data_Negative/" + station_name + "2023" + ".xml"
    tree = ET.parse(rainfile)
    root = tree.getroot()

    if len(root[0][0]) != 0:
        if isinstance(root[0][0][day-2][0].text, type(None)):
            day0 = 0
        else:
            day0 = float(root[0][0][day-2][0].text)

        if isinstance(root[0][0][day-1][0].text, type(None)):
            day1 = 0
        else:
            day1 = float(root[0][0][day-1][0].text)
    else:
        if isinstance(root[0][4][day-2][0].text, type(None)):
            day0 = 0
        else:
            day0 = float(root[0][4][day-2][0].text)

        if isinstance(root[0][4][day-1][0].text, type(None)):
            day1 = 0
        else:
            day1 = float(root[0][4][day-1][0].text)
    ws['G' + str(row_index + 1)] = day0
    ws['H' + str(row_index + 1)] = day1

wb.save("modified" + file + ".xlsx")
