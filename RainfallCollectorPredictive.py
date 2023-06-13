from openpyxl.reader.excel import load_workbook
import xml.etree.ElementTree as ET

file = "rainfall_negative.xlsx"

wb = load_workbook(filename = file, data_only=True)
ws = wb.active

landslides = tuple(ws.rows)

for row_index in range(1, len(landslides)):
    entry = landslides[row_index]
    if str(entry[14].value) == "yes":
        station_name = str(entry[10].value)
        year1 = entry[16].value
        year2 = entry[17].value
        day = entry[18].value
        rainfile1 = "Rainfall_Data_Negative/" + station_name + str(year1) + ".xml"
        rainfile2 = "Rainfall_Data_Negative/" + station_name + str(year2) + ".xml"
        tree = ET.parse(rainfile1)
        root = tree.getroot()
        print(root[0][0][2].attrib['date'])

        if len(root[0][0]) != 0:
            if isinstance(root[0][0][day-2][0].text, type(None)):
                year1day0 = 0
            else:
                year1day0 = float(root[0][0][day-2][0].text)

            if isinstance(root[0][0][day-1][0].text, type(None)):
                year1day1 = 0
            else:
                year1day1 = float(root[0][0][day-1][0].text)
        else:
            year1day0 = 0
            year1day1 = 0

        tree = ET.parse(rainfile2)
        root = tree.getroot()
        if len(root[0][0]) != 0:
            if isinstance(root[0][0][day - 2][0].text, type(None)):
                year2day0 = 0
            else:
                year2day0 = float(root[0][0][day - 2][0].text)

            if isinstance(root[0][0][day - 1][0].text, type(None)):
                year2day1 = 0
            else:
                year2day1 = float(root[0][0][day - 1][0].text)
        else:
            year2day0 = 0
            year2day1 = 0
    ws['T' + str(row_index + 1)] = year1day0
    ws['U' + str(row_index + 1)] = year1day1
    ws['V' + str(row_index + 1)] = year2day0
    ws['W' + str(row_index + 1)] = year2day1

wb.save("modified" + file + ".xlsx")
