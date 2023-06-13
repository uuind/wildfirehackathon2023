from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from requests import get
import json

from haversine import haversine

file = input("Please enter filename of data");
workbook_data = load_workbook(filename = file);
print("File Loaded!")
sheet = input("Please enter specific sheet name");
workbook = workbook_data[sheet]

latitude_col = 0
longitude_col = 0

location_data = 'WeatherStationLocations.json'
with open(location_data, "r") as read_file:
    all_stations = json.load(read_file)

def find_closest(my_lat, my_lon):
    smallest = 12450
    for station in all_stations:
        station_lat = station['Latitude']
        station_long = station['Longitude']
        distance = haversine(my_lat, my_lon, station_lat, station_long)
        if distance < smallest:
            smallest = distance
            closest_station = station['Station Location']
    return closest_station

storage = tuple(workbook.rows)

for col in storage[0]:
    if col.value.casefold() == 'latitude':
        latitude_col = col.col_idx
    elif col.value.casefold() == 'longitude':
        longitude_col = col.col_idx

location_col = longitude_col + 1
location_letter = get_column_letter(location_col)
workbook.insert_cols(location_col)
workbook[location_letter + "1"] = "weather_station"


for row_index in range(1, len(storage)):
    row = storage[row_index]
    input_lat = float(row[latitude_col - 1].value)
    input_long = float(row[longitude_col - 1].value)
    workbook[location_letter + str(row_index+1)] = find_closest(input_lat, input_long)

workbook_data.save("modified" + file)