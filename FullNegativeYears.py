import os
from copy import copy
import xml.etree.ElementTree as ET
from openpyxl.reader.excel import load_workbook

file = "modifiedWildfireCheck.xlsx"

wb = load_workbook(filename = file, data_only=True)
ws = wb.active
ws2 = wb.create_sheet("test")

wildfires = tuple(ws.rows)

year_check = [2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023]
years = []
overall_index = 0
for row_index in range(1, len(wildfires)-1):
    entry = wildfires[row_index]
    next_entry = wildfires[row_index+1]
    temp = []
    if entry[5].value not in years:
        years.append(entry[5].value)
    if entry[6].value not in years:
        years.append(entry[6].value)
    if entry[0].value != next_entry[0].value:
        temp = copy(year_check)
        for check in years:
            temp.remove(check)

        print(entry[0].value)
        print(temp)
        years = []
    for year in temp:
        if year != 2007:
            rainfile = "Rainfall_Data_Negative/" + entry[0].value + str(year) + ".xml"
            if os.path.exists(rainfile):
                tree = ET.parse(rainfile)
                root = tree.getroot()
                if len(root[0][0]) != 0:
                    for i in range(1, len(root[0][0])):
                        if isinstance(root[0][0][i-1][0].text, type(None)):
                            day = 0
                        else:
                            day = float(root[0][0][i-1][0].text)
                        ws2['A' + str(overall_index + 1)] = entry[0].value
                        ws2['B' + str(overall_index + 1)] = str(root[0][0][i-1].attrib['date'])
                        ws2['C' + str(overall_index + 1)] = day
                        overall_index += 1
                    else:
                        day = 0
                else:
                    for i2 in range(1, len(root[0][4])):
                        if isinstance(root[0][4][i2 - 1][0].text, type(None)):
                            day = 0
                        else:
                            day = float(root[0][4][i2 - 1][0].text)
                        ws2['A' + str(overall_index + 1)] = entry[0].value
                        ws2['B' + str(overall_index + 1)] = str(root[0][4][i2 - 1].attrib['val'])
                        ws2['C' + str(overall_index + 1)] = day
                        overall_index += 1
                    else:
                        day = 0

wb.save("fullynegativerainfall.xlsx")

